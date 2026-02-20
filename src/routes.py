
'''
- Public routes (landing, order form, status check)
- Admin routes (login, dashboard, management)
- Export functions (Excel, PDF) '''

from flask import render_template, request, redirect, url_for, flash, send_from_directory, session, jsonify, send_file
from werkzeug.utils import secure_filename
from models import db, Pesanan, get_statistics
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'unitproduksi123'

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'docx'}


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def register_routes(app):
    """
    Register semua routes ke Flask app
    
    Args:
        app: Flask application instance
    """
    
    # ==================== rute public ====================
    
    @app.route('/')
    def index():
        """Landing page dengan katalog produk"""
        return render_template('produk.html')
    
    @app.route('/pesan', methods=['GET', 'POST'])
    def pesan():
        """Form pemesanan"""
        if request.method == 'POST':
            nama = request.form.get('nama')
            kontak = request.form.get('kontak')
            jenis_print = request.form.get('jenis_print')
            ukuran = request.form.get('ukuran')
            jumlah = request.form.get('jumlah')
            
            if not all([nama, kontak, jenis_print, ukuran, jumlah]):
                flash('Semua field harus diisi!', 'danger')
                return redirect(url_for('pesan'))
            
            file_path = None
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{timestamp}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(file_path)
                    file_path = filename  
            
            try:
                pesanan_baru = Pesanan(
                    nama=nama,
                    kontak=kontak,
                    jenis_print=jenis_print,
                    ukuran=ukuran,
                    jumlah=int(jumlah),
                    file_path=file_path,
                    status='pending'
                )
                
                db.session.add(pesanan_baru)
                db.session.commit()
                
                flash(f'Pesanan berhasil dibuat! Nomor pesanan Anda: #{pesanan_baru.id}', 'success')
                return redirect(url_for('cek_status'))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Terjadi kesalahan: {str(e)}', 'danger')
                return redirect(url_for('pesan'))
        
        return render_template('user/index.html')
    
    @app.route('/cek-status', methods=['GET', 'POST'])
    def cek_status():
        """Check order status"""
        pesanan = None
        
        if request.method == 'POST':
            nomor_pesanan = request.form.get('nomor_pesanan')
            
            if nomor_pesanan:
                try:
                    pesanan_id = int(nomor_pesanan.replace('#', ''))
                    pesanan = Pesanan.query.get(pesanan_id)
                    
                    if not pesanan:
                        flash('Nomor pesanan tidak ditemukan!', 'warning')
                except ValueError:
                    flash('Format nomor pesanan salah!', 'danger')
        
        return render_template('user/cek_status.html', pesanan=pesanan)
    
    @app.route('/uploads/<filename>')
    def uploaded_file(filename):
        """Serve uploaded files"""
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    
    # ==================== rute ke admin ====================
    
    @app.route('/admin/login', methods=['GET', 'POST'])
    def admin_login():
        """Admin login page"""
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            
            if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                session['admin_logged_in'] = True
                flash('Login berhasil!', 'success')
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Username atau password salah!', 'danger')
        
        return render_template('admin/login.html')
    
    @app.route('/admin/logout')
    def admin_logout():
        """Admin logout"""
        session.pop('admin_logged_in', None)
        flash('Logout berhasil!', 'info')
        return redirect(url_for('admin_login'))
    
    @app.route('/admin/dashboard')
    def admin_dashboard():
        """Admin dashboard - requires login"""
        if not session.get('admin_logged_in'):
            flash('Silakan login terlebih dahulu!', 'warning')
            return redirect(url_for('admin_login'))
        
        status_filter = request.args.get('status', 'all')
        search_query = request.args.get('search', '')
        
        query = Pesanan.query
        
        if status_filter != 'all':
            query = query.filter_by(status=status_filter)
        
        if search_query:
            query = query.filter(
                db.or_(
                    Pesanan.nama.ilike(f'%{search_query}%'),
                    Pesanan.kontak.ilike(f'%{search_query}%')
                )
            )
        
        pesanan_list = query.order_by(Pesanan.created_at.desc()).all()
        
        stats = get_statistics()
        
        return render_template(
            'admin/dashboard.html',
            pesanan_list=pesanan_list,
            stats=stats,
            current_filter=status_filter,
            search_query=search_query
        )
    
    @app.route('/admin/update-status/<int:id>', methods=['POST'])
    def update_status(id):
        """Update order status"""
        if not session.get('admin_logged_in'):
            return jsonify({'error': 'Unauthorized'}), 401
        
        pesanan = Pesanan.query.get_or_404(id)
        new_status = request.form.get('status')
        
        if new_status in ['pending', 'proses', 'selesai']:
            pesanan.status = new_status
            db.session.commit()
            flash('Status pesanan berhasil diupdate!', 'success')
        else:
            flash('Status tidak valid!', 'danger')
        
        return redirect(url_for('admin_dashboard'))
    
    @app.route('/admin/delete/<int:id>', methods=['POST'])
    def delete_pesanan(id):
        """Delete order"""
        if not session.get('admin_logged_in'):
            return jsonify({'error': 'Unauthorized'}), 401
        
        pesanan = Pesanan.query.get_or_404(id)
        
        if pesanan.file_path:
            file_full_path = os.path.join(app.config['UPLOAD_FOLDER'], pesanan.file_path)
            if os.path.exists(file_full_path):
                os.remove(file_full_path)
        
        db.session.delete(pesanan)
        db.session.commit()
        
        flash('Pesanan berhasil dihapus!', 'success')
        return redirect(url_for('admin_dashboard'))
    
    
    @app.route('/admin/api/stats')
    def api_stats():
        """API endpoint untuk statistik (AJAX)"""
        if not session.get('admin_logged_in'):
            return jsonify({'error': 'Unauthorized'}), 401
        
        stats = get_statistics()
        return jsonify(stats)
    
    # ==================== EXPORT ROUTES ====================
    
    @app.route('/admin/export/excel')
    def export_excel():
        """Export pesanan ke Excel"""
        if not session.get('admin_logged_in'):
            flash('Silakan login terlebih dahulu!', 'warning')
            return redirect(url_for('admin_login'))
        
        pesanan_list = Pesanan.query.order_by(Pesanan.created_at.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Pesanan"
        
        headers = ['No. Pesanan', 'Nama', 'Kontak', 'Jenis Print', 'Ukuran', 'Jumlah', 'Status', 'Tanggal']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, pesanan in enumerate(pesanan_list, 2):
            ws.cell(row=row_idx, column=1).value = f"#{pesanan.id}"
            ws.cell(row=row_idx, column=2).value = pesanan.nama
            ws.cell(row=row_idx, column=3).value = pesanan.kontak
            ws.cell(row=row_idx, column=4).value = pesanan.jenis_print
            ws.cell(row=row_idx, column=5).value = pesanan.ukuran
            ws.cell(row=row_idx, column=6).value = pesanan.jumlah
            ws.cell(row=row_idx, column=7).value = pesanan.status_indo
            ws.cell(row=row_idx, column=8).value = pesanan.created_at.strftime('%Y-%m-%d %H:%M')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Data_Pesanan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    @app.route('/admin/export/pdf')
    def export_pdf():
        """Export pesanan ke PDF"""
        if not session.get('admin_logged_in'):
            flash('Silakan login terlebih dahulu!', 'warning')
            return redirect(url_for('admin_login'))
        
        pesanan_list = Pesanan.query.order_by(Pesanan.created_at.desc()).all()
        
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, height - 50, "LAPORAN DATA PESANAN")
        c.drawString(50, height - 70, "Fikri Production - Unit Produksi SMKIT Ihsanul Fikri")
        
        c.setFont("Helvetica", 10)
        c.drawString(50, height - 90, f"Tanggal Export: {datetime.now().strftime('%d %B %Y %H:%M')}")
        
        y = height - 130
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "No.")
        c.drawString(100, y, "Nama")
        c.drawString(250, y, "Kontak")
        c.drawString(350, y, "Jenis Print")
        c.drawString(480, y, "Status")
        
        y -= 5
        c.line(50, y, width - 50, y)
        
        c.setFont("Helvetica", 9)
        y -= 20
        
        for pesanan in pesanan_list:
            if y < 50:  
                c.showPage()
                y = height - 50
                c.setFont("Helvetica", 9)
            
            c.drawString(50, y, f"#{pesanan.id}")
            c.drawString(100, y, pesanan.nama[:18])  
            c.drawString(250, y, pesanan.kontak[:12])
            c.drawString(350, y, pesanan.jenis_print[:15])
            c.drawString(480, y, pesanan.status_indo)
            
            y -= 20
        
        c.save()
        buffer.seek(0)
        
        filename = f"Laporan_Pesanan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
